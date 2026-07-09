from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
import pytest

from app.schemas.labor import LaborInputRow
from app.services.excel_reader import read_labor_rows
from app.services.excel_writer import build_template_workbook
from app.services.test_workbooks import (
    build_error_test_workbook,
    build_logic_test_workbook,
    build_result_workbook,
)


def make_row(amount=3000, name="张三", id_no="110101199001011234", year=2026, month=6):
    return LaborInputRow(
        year=year,
        month=month,
        department="事业部A",
        province="北京",
        reimburser="报销人",
        accountant="会计",
        name=name,
        id_no=id_no,
        after_tax_amount=Decimal(str(amount)),
    )


def has_no_fill(cell) -> bool:
    return cell.fill.fill_type is None


def test_template_can_be_read_by_reader():
    data = build_template_workbook()
    rows = read_labor_rows(BytesIO(data))
    assert len(rows) == 1
    assert rows[0].year == 2026
    assert rows[0].month == 6
    assert rows[0].name == "张三"
    assert rows[0].after_tax_amount == Decimal("3000")


def test_logic_test_template_has_expected_result_sheet_and_covers_key_scenarios():
    data = build_logic_test_workbook()
    wb = load_workbook(BytesIO(data), data_only=True)
    assert wb.sheetnames == ["逻辑测试数据", "预期结果说明", "说明"]
    expected_ws = wb["预期结果说明"]
    assert expected_ws["A1"].value == "场景编号"
    assert expected_ws["A2"].value == "T01"
    assert "跨800" in expected_ws["B3"].value

    rows = read_labor_rows(BytesIO(data))
    assert len(rows) == 21

    names = [row.name for row in rows]
    assert "测试A-免税累计" in names
    assert "测试B-跨800" in names
    assert "测试F-超过49500" in names
    assert "测试G-增值税临界500" in names
    assert "测试H-增值税超过500" in names

    cross_month = [row for row in rows if row.id_no == "110101199001010009"]
    assert [row.month for row in cross_month] == [6, 7]

    cross_year = [row for row in rows if row.id_no == "110101199001010010"]
    assert [row.year for row in cross_year] == [2026, 2027]

    same_name_rows = [row for row in rows if row.name == "张三"]
    assert len({row.id_no for row in same_name_rows}) == 2


def test_error_test_template_reports_all_invalid_rows_with_readable_messages():
    data = build_error_test_workbook()
    with pytest.raises(ValueError) as exc_info:
        read_labor_rows(BytesIO(data))
    message = str(exc_info.value)
    for row_number in range(2, 11):
        assert f"第 {row_number} 行" in message
    assert "年份" in message and "不能为空" in message
    assert "月份" in message and "不能为空" in message
    assert "月份必须" in message and "1-12" in message
    assert "身份证号码不能为空" in message
    assert "姓名不能为空" in message
    assert "税后劳务金额不能为空" in message
    assert "税后劳务金额必须大于" in message
    assert "税后劳务金额必须是数字" in message


def test_result_workbook_has_value_sheets_formula_sheet_and_test_report():
    data = build_result_workbook([make_row(400), make_row(500)])
    wb = load_workbook(BytesIO(data), data_only=False)
    assert wb.sheetnames == ["劳务费税费换算台账", "清晰版台账", "公式版台账", "测试核对报告"]

    ws1 = wb["劳务费税费换算台账"]
    assert ws1["A1"].value is None
    assert ws1["J1"].value is None
    assert has_no_fill(ws1["A1"])
    assert has_no_fill(ws1["J1"])
    assert ws1["A2"].value == "劳务费税前（后）相关税费换算台账"
    assert ws1["J4"].value == "税后劳务金额"
    assert ws1["X4"].value == "核对"
    assert ws1["A6"].value == 1
    assert ws1["J6"].value == Decimal("400.00")
    assert has_no_fill(ws1["A6"])
    assert has_no_fill(ws1["J6"])
    assert ws1["K7"].value == Decimal("900.00")
    assert not isinstance(ws1["L7"].value, str) or not ws1["L7"].value.startswith("=")

    ws2 = wb["清晰版台账"]
    assert ws2["A1"].value == "基础信息"
    assert ws2["W2"].value == "核对结果"

    ws3 = wb["公式版台账"]
    assert ws3["A1"].value is None
    assert ws3["J1"].value is None
    assert has_no_fill(ws3["A1"])
    assert has_no_fill(ws3["J1"])
    assert ws3["A2"].value == "劳务费税前（后）相关税费换算台账"
    assert ws3["A6"].value == 1
    assert ws3["J6"].value == Decimal("400.00")
    assert has_no_fill(ws3["A6"])
    assert has_no_fill(ws3["J6"])
    assert ws3["K7"].value.startswith("=SUMIFS(")
    assert ws3["L7"].value.startswith("=IF(")
    assert ws3["Q7"].value.startswith("=IFERROR(")
    assert ws3["S7"].value.startswith("=MAX(0,")
    assert ws3["X7"].value == "=J7+Q7+R7-O7"

    ws4 = wb["测试核对报告"]
    assert ws4["A1"].value == "场景编号"
    assert ws4["A2"].value == "基础核对"
    assert ws4["F2"].value == "通过"


def test_logic_template_export_generates_passing_test_report():
    logic_rows = read_labor_rows(BytesIO(build_logic_test_workbook()))
    data = build_result_workbook(logic_rows)
    wb = load_workbook(BytesIO(data), data_only=True)
    ws = wb["测试核对报告"]
    statuses = [ws.cell(row_idx, 6).value for row_idx in range(2, ws.max_row + 1)]
    assert statuses
    assert set(statuses) == {"通过"}
    scenario_ids = [ws.cell(row_idx, 1).value for row_idx in range(2, ws.max_row + 1)]
    for scenario_id in ["T01", "T02", "T03", "T04", "T05", "T06", "T07", "T08", "T09", "T10", "T11"]:
        assert scenario_id in scenario_ids
