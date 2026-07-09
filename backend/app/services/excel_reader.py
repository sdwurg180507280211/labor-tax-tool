from __future__ import annotations

from datetime import date, datetime
from typing import Any, BinaryIO
import re

from openpyxl import load_workbook
from pydantic import ValidationError

from app.schemas.labor import LaborInputRow

HEADER_ALIASES: dict[str, str] = {
    "会议日期": "event_date",
    "事业部": "department",
    "发起人": "reimburser",
    "讲者姓名": "name",
    "讲者id": "id_no",
    "劳务费（实付金额）": "after_tax_amount",
    "劳务费(实付金额)": "after_tax_amount",
    "劳务费实付金额": "after_tax_amount",
}

REQUIRED_FIELDS = {"event_date", "name", "id_no", "after_tax_amount"}
FIELD_LABELS = {
    "event_date": "会议日期",
    "name": "讲者姓名",
    "id_no": "讲者ID",
    "after_tax_amount": "劳务费（实付金额）",
}


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value)
    return (
        text.strip()
        .replace(" ", "")
        .replace("\n", "")
        .replace("\r", "")
        .replace("_", "")
        .lower()
    )


def find_header_row(ws) -> tuple[int, dict[int, str]]:
    """Find the first row that matches the latest backend export template."""
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        mapping: dict[int, str] = {}
        for cell in ws[row_idx]:
            normalized = normalize_header(cell.value)
            if normalized in HEADER_ALIASES:
                mapping[cell.column] = HEADER_ALIASES[normalized]
        if len(set(mapping.values())) >= 3:
            return row_idx, mapping
    raise ValueError("未识别到后台导出表头。请使用最新模板，需包含：会议日期、讲者姓名、讲者ID、劳务费（实付金额）。")


def _parse_event_year_month(value: Any) -> tuple[int, int]:
    if value in (None, ""):
        raise ValueError("会议日期不能为空")
    if isinstance(value, datetime):
        return value.year, value.month
    if isinstance(value, date):
        return value.year, value.month

    text = str(value).strip()
    if not text:
        raise ValueError("会议日期不能为空")

    patterns = [
        r"^(\d{4})[-/\.](\d{1,2})[-/\.](\d{1,2})$",
        r"^(\d{4})年(\d{1,2})月(\d{1,2})日$",
        r"^(\d{4})(\d{2})(\d{2})$",
    ]
    for pattern in patterns:
        match = re.match(pattern, text)
        if match:
            year = int(match.group(1))
            month = int(match.group(2))
            if not 1 <= month <= 12:
                raise ValueError("会议日期月份必须在1-12之间")
            return year, month
    raise ValueError("会议日期格式不正确，请使用日期格式，如 2026-06-01")


def _format_validation_error(err: dict[str, Any]) -> str:
    field = str(err.get("loc", [""])[0])
    label = FIELD_LABELS.get(field, field)
    message = str(err.get("msg", "校验失败"))
    error_type = str(err.get("type", ""))

    if field == "year":
        return "会议日期年份必须在1900-2999之间"
    if field == "month":
        return "会议日期月份必须在1-12之间"
    if field == "name":
        return "讲者姓名不能为空"
    if field == "id_no":
        return "讲者ID不能为空"
    if field == "after_tax_amount":
        if "不能为空" in message:
            return "劳务费（实付金额）不能为空"
        if "必须是数字" in message:
            return "劳务费（实付金额）必须是数字"
        if "greater than" in message:
            return "劳务费（实付金额）必须大于0"
        return "劳务费（实付金额）格式不正确"
    return f"{label}：{message or error_type}"


def read_labor_rows(file_obj: BinaryIO) -> list[LaborInputRow]:
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    header_row, col_to_field = find_header_row(ws)
    found_fields = set(col_to_field.values())
    missing = REQUIRED_FIELDS - found_fields
    if missing:
        raise ValueError("缺少必填列：" + "、".join(FIELD_LABELS[x] for x in sorted(missing)))

    rows: list[LaborInputRow] = []
    errors: list[str] = []
    for excel_row_idx in range(header_row + 1, ws.max_row + 1):
        raw: dict[str, Any] = {}
        event_date_value: Any = None
        empty = True
        for col_idx, field in col_to_field.items():
            value = ws.cell(excel_row_idx, col_idx).value
            if value not in (None, ""):
                empty = False
            if field == "event_date":
                event_date_value = value
            else:
                raw[field] = value
        if empty:
            continue
        try:
            year, month = _parse_event_year_month(event_date_value)
            raw["year"] = year
            raw["month"] = month
            rows.append(LaborInputRow(**raw))
        except ValueError as exc:
            errors.append(f"第 {excel_row_idx} 行：{exc}")
        except ValidationError as exc:
            msg = "；".join(_format_validation_error(err) for err in exc.errors())
            errors.append(f"第 {excel_row_idx} 行：{msg}")

    if errors:
        raise ValueError("\n".join(errors))
    if not rows:
        raise ValueError("没有读取到有效数据行。")
    return rows
