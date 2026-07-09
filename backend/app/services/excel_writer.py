from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.schemas.labor import LaborCalculatedRow, LaborInputRow
from app.services.tax_calculator import calculate_rows, money2_decimal

BACKEND_IMPORT_HEADERS = [
    "会议ID",
    "会议类型",
    "事业部",
    "领域",
    "产品",
    "发起人",
    "发起人员工ID",
    "主题",
    "会议日期",
    "会议开始时间",
    "会议结束时间",
    "有效听众人数",
    "在库听众数",
    "会议是否有效",
    "（讲者DEMS编码）",
    "讲者姓名",
    "讲者ID",
    "客户级别",
    "劳务角色",
    "讲者医院",
    "讲者科室",
    "劳务费（实付金额）",
    "个税",
    "劳务费（合计）",
]
TEMPLATE_HEADERS = BACKEND_IMPORT_HEADERS

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
SUBHEADER_FILL = PatternFill("solid", fgColor="EAF4FB")
THIN = Side(style="thin", color="000000")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
MONEY_FORMAT = '#,##0.00'
INTEGER_FORMAT = '0'
TEXT_FORMAT = '@'


def _save_workbook(wb: Workbook) -> bytes:
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()


def _set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _style_range(ws, row_start: int, row_end: int, col_start: int, col_end: int) -> None:
    for row in ws.iter_rows(min_row=row_start, max_row=row_end, min_col=col_start, max_col=col_end):
        for cell in row:
            cell.border = BORDER_THIN
            cell.alignment = CENTER


def _money(value: Decimal) -> Decimal:
    return money2_decimal(value)


def _backend_sample_row() -> list[object]:
    return [
        "M0001", "线下会", "测试事业部", "慢病管理", "测试产品", "测试发起人", "E0001", "测试会议",
        "2026-06-01", "09:00", "10:00", 30, 28, "是", "D0001", "张三", "SPK000001", "A",
        "讲者", "测试医院", "心内科", 3000, None, None,
    ]


def build_template_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    ws.append(BACKEND_IMPORT_HEADERS)
    ws.append(_backend_sample_row())
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(BACKEND_IMPORT_HEADERS)):
        for cell in row:
            cell.border = BORDER_THIN
            cell.alignment = LEFT if cell.column in (8, 16, 20, 21, 24) else CENTER
            if cell.row == 1:
                cell.fill = HEADER_FILL
                cell.font = Font(bold=True)
            if cell.column == 17:
                cell.number_format = TEXT_FORMAT
            if cell.column in (22, 23, 24):
                cell.number_format = MONEY_FORMAT
    _set_widths(ws, {"A": 12, "B": 12, "C": 14, "D": 12, "E": 14, "F": 12, "G": 16, "H": 26, "I": 14, "J": 14, "K": 14, "L": 16, "M": 14, "N": 14, "O": 18, "P": 14, "Q": 20, "R": 12, "S": 12, "T": 24, "U": 14, "V": 18, "W": 14, "X": 16})
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:X{ws.max_row}"
    return _save_workbook(wb)


def build_result_workbook(input_rows: list[LaborInputRow]) -> bytes:
    calculated_rows = calculate_rows(input_rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "劳务费税费换算台账"
    _write_simplified_ledger_sheet(ws, calculated_rows)
    _write_rule_sheet(wb.create_sheet("规则说明"))
    return _save_workbook(wb)


def _write_simplified_ledger_sheet(ws, rows: Iterable[LaborCalculatedRow]) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:N1")
    ws["A1"] = "劳务费税前（后）相关税费换算台账"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = CENTER
    ws["M3"] = "单位：元"
    ws["M3"].alignment = CENTER
    for merged in ["A4:A5", "B4:B5", "C4:C5", "D4:D5", "E4:E5", "F4:F5", "G4:H4", "I4:I5", "J4:J5", "K4:K5", "L4:N4"]:
        ws.merge_cells(merged)
    for coord, value in {
        "A4": "序号", "B4": "年份", "C4": "月份", "D4": "日期", "E4": "姓名", "F4": "讲者ID", "G4": "税后劳务金额",
        "I4": "累计税前金额\n-含个税不含增值税", "J4": "单日累计税前金额\n-含个税不含增值税", "K4": "单次税前金额\n-含个税、增值税", "L4": "单次代扣代缴税额",
        "G5": "单次", "H5": "本月累计", "L5": "增值税", "M5": "附加税", "N5": "个税",
    }.items():
        ws[coord] = value
    _style_range(ws, 4, 5, 1, 14)
    for row in ws.iter_rows(min_row=4, max_row=5, min_col=1, max_col=14):
        for cell in row:
            cell.fill = HEADER_FILL if cell.row == 4 else SUBHEADER_FILL
            cell.font = Font(bold=True)
    _set_widths(ws, {"A": 6, "B": 10, "C": 8, "D": 8, "E": 14, "F": 20, "G": 16, "H": 16, "I": 24, "J": 24, "K": 24, "L": 14, "M": 14, "N": 14})
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[4].height = 34
    ws.row_dimensions[5].height = 22
    row_idx = 6
    for item in rows:
        values = [item.index, item.year, item.month, item.day, item.name, item.id_no, _money(item.after_tax_amount), _money(item.cumulative_after_tax_amount), _money(item.cumulative_pre_tax_without_vat), _money(item.daily_cumulative_pre_tax_without_vat), _money(item.contract_amount), _money(item.vat_amount), _money(item.surcharge_amount), _money(item.individual_tax_amount)]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = BORDER_THIN
            cell.alignment = LEFT if col_idx in (5, 6) else CENTER
            if col_idx in (1, 2, 3, 4):
                cell.number_format = INTEGER_FORMAT
            elif col_idx == 6:
                cell.number_format = TEXT_FORMAT
            elif col_idx >= 7:
                cell.number_format = MONEY_FORMAT
        row_idx += 1
    if row_idx == 6:
        ws.cell(6, 1, "无数据")
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:N{max(row_idx - 1, 6)}"


def _write_rule_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    rows = [
        ["规则项", "新版0709口径"],
        ["导入来源", "后台导出表格.xlsx：读取会议日期、讲者姓名、讲者ID、劳务费（实付金额）"],
        ["个税累计维度", "年份 + 月份 + 讲者ID"],
        ["增值税累计维度", "年份 + 月份 + 日期 + 讲者ID"],
        ["本月累计税后", "同一讲者ID同年月税后劳务金额累计"],
        ["累计税前反推", "≤800：税前=税后；≤3360：(税后-160)/0.8；≤21000：税后/0.84；≤49500：(税后-2000)/0.76；>49500：(税后-7000)/0.68"],
        ["单日累计税前", "同一讲者ID同日的税后劳务金额累计 + 同日个税累计"],
        ["增值税", "单日累计税前≤1000：0；单日累计税前>1000：单日累计税前×1%"],
        ["附加税", "增值税 × 12% × 50%，即增值税×6%"],
        ["协议签订金额", "本次协议签订金额 = 当前单日累计税前+当前增值税+当前附加税 - 之前同日已确认协议签订金额"],
        ["个税", "按本月累计税前计算累计个税，再减去之前同年月已扣个税，得到本次个税"],
    ]
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = BORDER_THIN
            cell.alignment = LEFT
            if cell.row == 1:
                cell.fill = HEADER_FILL
                cell.font = Font(bold=True)
    _set_widths(ws, {"A": 22, "B": 110})
    ws.freeze_panes = "A2"
