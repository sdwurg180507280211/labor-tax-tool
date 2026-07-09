from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.schemas.labor import LaborCalculatedRow, LaborInputRow
from app.services.excel_writer import build_result_workbook as build_base_result_workbook
from app.services.tax_calculator import calculate_rows, money2_decimal

BACKEND_IMPORT_HEADERS = [
    "会议ID",
    "会议类型",
    "事业部",
    "领域",
    "产品",
    "发起人",
    "发起人员工ID",
    "主题",
    "会议日期",
    "会议开始时间",
    "会议结束时间",
    "有效听众人数",
    "在库听众数",
    "会议是否有效",
    "（讲者DEMS编码）",
    "讲者姓名",
    "讲者ID",
    "客户级别",
    "劳务角色",
    "讲者医院",
    "讲者科室",
    "劳务费（实付金额）",
    "个税",
    "劳务费（合计）",
]


def _backend_row(
    index: int,
    name: str,
    speaker_id: str,
    amount: int | float | str,
    year: int = 2026,
    month: int = 6,
    day: int = 15,
) -> list[object]:
    return [
        f"M{index:04d}",
        "线下会",
        "测试事业部",
        "慢病管理",
        "测试产品",
        "测试报销人",
        "E0001",
        f"测试会议{index:02d}",
        f"{year}-{month:02d}-{day:02d}",
        "09:00",
        "10:00",
        30,
        28,
        "是",
        f"D{index:04d}",
        name,
        speaker_id,
        "A",
        "讲者",
        "测试医院",
        "心内科",
        amount,
        None,
        None,
    ]


LOGIC_TEST_ROWS = [
    _backend_row(1, "测试A-免税累计", "110101199001010001", 400),
    _backend_row(2, "测试A-免税累计", "110101199001010001", 300),
    _backend_row(3, "测试B-跨800", "110101199001010002", 400),
    _backend_row(4, "测试B-跨800", "110101199001010002", 500),
    _backend_row(5, "测试C-3360内", "110101199001010003", 3000),
    _backend_row(6, "测试D-21000内", "110101199001010004", 5000),
    _backend_row(7, "测试D-21000内", "110101199001010004", 10000),
    _backend_row(8, "测试E-49500内", "110101199001010005", 22000),
    _backend_row(9, "测试E-49500内", "110101199001010005", 10000),
    _backend_row(10, "测试F-超过49500", "110101199001010006", 50000),
    _backend_row(11, "测试F-超过49500", "110101199001010006", 10000),
    _backend_row(12, "测试G-增值税临界500", "110101199001010007", 500),
    _backend_row(13, "测试H-增值税超过500", "110101199001010008", 501),
    _backend_row(14, "测试I-跨月", "110101199001010009", 3000, 2026, 6),
    _backend_row(15, "测试I-跨月", "110101199001010009", 3000, 2026, 7),
    _backend_row(16, "测试J-跨年", "110101199001010010", 3000, 2026, 6),
    _backend_row(17, "测试J-跨年", "110101199001010010", 3000, 2027, 6),
    _backend_row(18, "张三", "110101199001010011", 3000),
    _backend_row(19, "张三", "110101199001010012", 3000),
    _backend_row(20, "测试K-不同人同月", "110101199001010013", 1800),
    _backend_row(21, "测试L-不同人同月", "110101199001010014", 1800),
]

ERROR_TEST_ROWS = [
    _backend_row(1, "缺日期", "110101199001019001", 3000, 2026, 6),
    _backend_row(2, "日期格式错", "110101199001019002", 3000, 2026, 6),
    _backend_row(3, "缺讲者姓名", "110101199001019003", 3000, 2026, 6),
    _backend_row(4, "缺讲者ID", "", 3000, 2026, 6),
    _backend_row(5, "金额为空", "110101199001019005", None, 2026, 6),
    _backend_row(6, "金额为零", "110101199001019006", 0, 2026, 6),
    _backend_row(7, "金额为负", "110101199001019007", -100, 2026, 6),
    _backend_row(8, "金额文本", "110101199001019008", "abc", 2026, 6),
]
ERROR_TEST_ROWS[0][8] = None
ERROR_TEST_ROWS[1][8] = "2026-13-01"
ERROR_TEST_ROWS[2][15] = None
ERROR_TEST_ROWS[3][16] = None

EXPECTED_ROWS = [
    ["T01", "累计税后≤800", "测试A第2笔累计税后=700，个税=0", "验证免税区间"],
    ["T02", "同人同月跨800", "测试B第2笔累计税后=900，单次个税=25", "验证累计反推和个税拆分"],
    ["T03", "累计税后≤3360", "测试C累计税前=3550", "验证（税后-160）÷0.8"],
    ["T04", "累计税后3360-21000", "测试D第2笔累计税前=17857.14", "验证税后÷0.84"],
    ["T05", "累计税后21000-49500", "测试E第2笔累计税前=39473.68", "验证（税后-2000）÷0.76"],
    ["T06", "累计税后＞49500", "测试F第2笔累计税前=77941.18", "验证（税后-7000）÷0.68"],
    ["T07", "增值税500起征点", "测试G增值税=0，测试H增值税=5.01", "验证≤500不征、＞500按1%"],
    ["T08", "同一讲者ID跨月份", "测试I 7月累计税后=3000", "验证月份不同重新累计"],
    ["T09", "同一讲者ID跨年份同月份", "测试J 2027年6月累计税后=3000", "验证年份不同重新累计"],
    ["T10", "同名不同讲者ID", "两条张三分别累计3000", "验证不按姓名累计"],
    ["T11", "不同讲者同月", "测试K/L分别累计1800", "验证不同讲者ID互不累计"],
]

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
PASS_FILL = PatternFill("solid", fgColor="E2F0D9")
ERROR_FILL = PatternFill("solid", fgColor="FCE4D6")
BORDER = Border(*(Side(style="thin", color="000000"),) * 4)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
MONEY_FORMAT = "#,##0.00"
TEXT_FORMAT = "@"


def _save(wb: Workbook) -> bytes:
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()


def _set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _style_table(ws, max_col: int) -> None:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = BORDER
            cell.alignment = LEFT if cell.column in (8, 16, 20, 21, 24) else CENTER
            if cell.row == 1:
                cell.fill = HEADER_FILL
                cell.font = Font(bold=True)


def _clear_fill(cell) -> None:
    cell.fill = PatternFill(fill_type=None)


def _remove_export_input_highlights(wb: Workbook) -> None:
    """Remove yellow/manual-entry hints from exported result workbooks and use latest speaker labels."""
    for sheet_name in ("劳务费税费换算台账", "公式版台账"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for coord in ("A1", "J1"):
            ws[coord].value = None
            _clear_fill(ws[coord])
        ws["H4"] = "讲者姓名"
        ws["I4"] = "讲者ID"
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=10):
            for cell in row:
                _clear_fill(cell)
    if "清晰版台账" in wb.sheetnames:
        ws = wb["清晰版台账"]
        ws["H2"] = "讲者姓名"
        ws["I2"] = "讲者ID"


def _build_input_workbook(title: str, rows: list[list[object]], note: str, expected_rows: list[list[str]] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.append(BACKEND_IMPORT_HEADERS)
    for row in rows:
        ws.append(row)
    _style_table(ws, len(BACKEND_IMPORT_HEADERS))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(BACKEND_IMPORT_HEADERS)):
        row[16].number_format = TEXT_FORMAT
        row[21].number_format = MONEY_FORMAT
        row[22].number_format = MONEY_FORMAT
        row[23].number_format = MONEY_FORMAT
    _set_widths(ws, {
        "A": 12, "B": 12, "C": 14, "D": 12, "E": 14, "F": 12, "G": 16, "H": 26,
        "I": 14, "J": 14, "K": 14, "L": 16, "M": 14, "N": 14, "O": 18, "P": 14,
        "Q": 20, "R": 12, "S": 12, "T": 24, "U": 14, "V": 18, "W": 14, "X": 16,
    })
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:X{ws.max_row}"

    if expected_rows:
        expected_ws = wb.create_sheet("预期结果说明")
        expected_ws.append(["场景编号", "场景名称", "预期结果", "验证目的"])
        for row in expected_rows:
            expected_ws.append(row)
        _style_table(expected_ws, 4)
        _set_widths(expected_ws, {"A": 12, "B": 24, "C": 50, "D": 36})
        expected_ws.freeze_panes = "A2"
        expected_ws.auto_filter.ref = f"A1:D{expected_ws.max_row}"

    note_ws = wb.create_sheet("说明")
    note_ws["A1"] = note
    note_ws["A1"].alignment = LEFT
    note_ws["A1"].font = Font(bold=True)
    note_ws.column_dimensions["A"].width = 100
    return _save(wb)


def build_template_workbook() -> bytes:
    return _build_input_workbook(
        "后台导出模板",
        [_backend_row(1, "张三", "SPK000001", 3000, 2026, 6)],
        "正式导入模板：以后台导出表格字段为准。系统使用会议日期拆分年份和月份，使用讲者ID作为唯一累计标识，使用劳务费（实付金额）作为税后金额入口。",
    )


def build_logic_test_workbook() -> bytes:
    return _build_input_workbook(
        "逻辑测试数据",
        LOGIC_TEST_ROWS,
        "逻辑测试模板：基于最新后台导出字段，覆盖免税、跨800、3360内、21000内、49500内、超过49500、增值税500临界、跨月、跨年、同名不同讲者ID、不同讲者同月等正常计算场景。",
        EXPECTED_ROWS,
    )


def build_error_test_workbook() -> bytes:
    return _build_input_workbook(
        "异常测试数据",
        ERROR_TEST_ROWS,
        "异常测试模板：基于最新后台导出字段，用于测试校验提示，故意包含会议日期/讲者姓名/讲者ID/劳务费为空、日期月份错误、金额为0、负数和文本等错误数据。",
    )


def _money(value: Decimal) -> str:
    return str(money2_decimal(value))


def _report_row(scenario_id: str, name: str, item: str, actual: str, expected: str, passed: bool, note: str) -> list[str]:
    return [scenario_id, name, item, actual, expected, "通过" if passed else "不通过", note]


def _find(rows: list[LaborCalculatedRow], speaker_id: str, year: int | None = None, month: int | None = None) -> list[LaborCalculatedRow]:
    return [row for row in rows if row.id_no == speaker_id and (year is None or row.year == year) and (month is None or row.month == month)]


def _is_logic_rows(rows: list[LaborCalculatedRow]) -> bool:
    return {str(row[16]) for row in LOGIC_TEST_ROWS}.issubset({row.id_no for row in rows})


def _money_check(rows: list[list[str]], scenario_id: str, name: str, item: str, actual: Decimal, expected: str, note: str) -> None:
    actual_text = _money(actual)
    rows.append(_report_row(scenario_id, name, item, actual_text, expected, actual_text == expected, note))


def _build_report_rows(rows: list[LaborCalculatedRow]) -> list[list[str]]:
    report: list[list[str]] = []
    bad_checks = [row for row in rows if abs(row.check_amount) > Decimal("0.0049")]
    report.append(_report_row("基础核对", "核对列检查", "核对值非0行数", f"{len(bad_checks)} 行", "0 行", len(bad_checks) == 0, "所有业务数据均执行此检查。"))
    if not _is_logic_rows(rows):
        report.append(_report_row("提示", "非内置逻辑测试模板", "专项场景检查", "未执行", "使用逻辑测试模板时自动执行", True, "当前数据不是内置逻辑测试模板，仅输出基础核对结果。"))
        return report

    a = _find(rows, "110101199001010001")
    _money_check(report, "T01", "累计税后≤800", "测试A第2笔累计税后", a[1].cumulative_after_tax_amount, "700.00", "免税区间累计税前应等于税后。")
    _money_check(report, "T01", "累计税后≤800", "测试A第2笔单次个税", a[1].individual_tax_amount, "0.00", "免税区间个税应为0。")
    b = _find(rows, "110101199001010002")
    _money_check(report, "T02", "同人同月跨800", "测试B第2笔累计税后", b[1].cumulative_after_tax_amount, "900.00", "同一讲者ID同年月累计。")
    _money_check(report, "T02", "同人同月跨800", "测试B第2笔单次个税", b[1].individual_tax_amount, "25.00", "累计税前925，累计个税25。")
    _money_check(report, "T03", "累计税后≤3360", "测试C累计税前", _find(rows, "110101199001010003")[0].cumulative_pre_tax_without_vat, "3550.00", "按（税后-160）÷0.8反推。")
    _money_check(report, "T04", "累计税后3360-21000", "测试D第2笔累计税前", _find(rows, "110101199001010004")[1].cumulative_pre_tax_without_vat, "17857.14", "按税后÷0.84反推。")
    _money_check(report, "T05", "累计税后21000-49500", "测试E第2笔累计税前", _find(rows, "110101199001010005")[1].cumulative_pre_tax_without_vat, "39473.68", "按（税后-2000）÷0.76反推。")
    _money_check(report, "T06", "累计税后＞49500", "测试F第2笔累计税前", _find(rows, "110101199001010006")[1].cumulative_pre_tax_without_vat, "77941.18", "按（税后-7000）÷0.68反推。")
    _money_check(report, "T07", "增值税500起征点", "测试G单次增值税", _find(rows, "110101199001010007")[0].vat_amount, "0.00", "本次对应税前≤500，不计增值税。")
    _money_check(report, "T07", "增值税500起征点", "测试H单次增值税", _find(rows, "110101199001010008")[0].vat_amount, "5.01", "本次对应税前>500，按1%计增值税。")
    _money_check(report, "T08", "同一讲者ID跨月份", "测试I 7月累计税后", _find(rows, "110101199001010009", 2026, 7)[0].cumulative_after_tax_amount, "3000.00", "月份不同应重新累计。")
    _money_check(report, "T09", "同一讲者ID跨年份同月份", "测试J 2027年6月累计税后", _find(rows, "110101199001010010", 2027, 6)[0].cumulative_after_tax_amount, "3000.00", "年份不同应重新累计。")
    zhang_rows = [row for row in rows if row.name == "张三"]
    zhang_ok = len(zhang_rows) == 2 and all(_money(row.cumulative_after_tax_amount) == "3000.00" for row in zhang_rows)
    report.append(_report_row("T10", "同名不同讲者ID", "两条张三是否分别累计", "是" if zhang_ok else "否", "是", zhang_ok, "同名人员不能按姓名合并累计。"))
    diff_ok = _money(_find(rows, "110101199001010013")[0].cumulative_after_tax_amount) == "1800.00" and _money(_find(rows, "110101199001010014")[0].cumulative_after_tax_amount) == "1800.00"
    report.append(_report_row("T11", "不同讲者同月", "不同讲者ID是否互不累计", "是" if diff_ok else "否", "是", diff_ok, "同月不同讲者ID应分别累计。"))
    return report


def _write_test_report_sheet(ws, rows: list[LaborCalculatedRow]) -> None:
    ws.sheet_view.showGridLines = False
    headers = ["场景编号", "场景名称", "检查项", "实际值", "预期值", "是否通过", "说明"]
    ws.append(headers)
    for row in _build_report_rows(rows):
        ws.append(row)
    _style_table(ws, len(headers))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
        for cell in row:
            cell.fill = PASS_FILL if cell.value == "通过" else ERROR_FILL
            cell.font = Font(bold=True)
    _set_widths(ws, {"A": 14, "B": 26, "C": 28, "D": 18, "E": 18, "F": 12, "G": 48})
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{ws.max_row}"


def build_result_workbook(input_rows: list[LaborInputRow]) -> bytes:
    wb = load_workbook(BytesIO(build_base_result_workbook(input_rows)))
    _remove_export_input_highlights(wb)
    if "测试核对报告" in wb.sheetnames:
        del wb["测试核对报告"]
    ws = wb.create_sheet("测试核对报告")
    _write_test_report_sheet(ws, calculate_rows(input_rows))
    return _save(wb)
